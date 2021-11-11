import argparse
import datetime as dt
import os
from typing import List

import openpyxl
import pandas as pd

from src.config.appConfig import loadAppConfig
from src.repos.gensMasterDataRepo import GensMasterRepo
from src.repos.schDataRepo import SchedulesRepo
from src.repos.smpDataRepo import SmpRepo
from src.services.gamsService import runGams
from src.typeDefs.schRow import ISchRow
from src.typeDefs.smpRow import ISmpRow

# read config file
appConf = loadAppConfig()
dbHost = appConf["dbHost"]
dbName = appConf["dbName"]
dbUname = appConf["dbUname"]
dbPass = appConf["dbPass"]
gamsExePath = appConf["gamsExePath"]
gamsCodePath = appConf["gamsCodePath"]
gamsLstPath = appConf["gamsLstPath"]
gamsExcelPath = appConf["gamsExcelPath"]

# make default target date as today
targetDt = dt.datetime.now()
targetDt = dt.datetime(targetDt.year, targetDt.month, targetDt.day)

# get target date from command line if present
parser = argparse.ArgumentParser()
parser.add_argument('--date', help='target Date')
parser.add_argument('--doff', help='Date offset')
parser.add_argument('--rev', help='revision number')
parser.add_argument('--single', action="store_true")
args = parser.parse_args()
targetDtStr = args.date
if not targetDtStr == None:
    targetDt = dt.datetime.strptime(targetDtStr, "%Y-%m-%d")
targetDtOffsetStr = args.doff
targetDtOffset = 0
if not targetDtOffsetStr == None:
    targetDtOffset = int(targetDtOffsetStr)
# add offset days to target date
targetDt = targetDt + dt.timedelta(days=targetDtOffset)

# flag that specifies to process only target revision and avoid processing revisions beyond this revision
isProcessOnlyOneRev = args.single

targetGujRevStr = args.rev
targetGujRev = None
if not targetGujRevStr == None:
    targetGujRev = int(targetGujRevStr)

# check - check if gams excel file exists
isGamsExcelPresent = os.path.isfile(gamsExcelPath)

if not isGamsExcelPresent:
    print("GAMS input excel not present...")
    exit(0)

# get the generators info from db
gensRepo = GensMasterRepo(
    dbHost, dbName, dbUname, dbPass)
gens = gensRepo.getGens()

gamsExcel = openpyxl.load_workbook(gamsExcelPath)

# write data to generators data sheet
# Data sheet column numbers
gensSheetName = "Data"
stationColNum = 3
vcColNum = 8
unitCapColNum = 11
tmColNum = 12
rUpColNum = 13
rDnColNum = 14
if not gensSheetName in gamsExcel.sheetnames:
    print('Generators data sheet does not exist in gams input excel file')
    exit(0)
gensSheet = gamsExcel[gensSheetName]
for gItr, g in enumerate(gens):
    gensSheet.cell(row=gItr+2, column=stationColNum).value = g["name"]
    gensSheet.cell(row=gItr+2, column=vcColNum).value = g["vcPu"]
    gensSheet.cell(row=gItr+2, column=unitCapColNum).value = g["capPu"]
    gensSheet.cell(row=gItr+2, column=tmColNum).value = g["tmPu"]
    gensSheet.cell(row=gItr+2, column=rUpColNum).value = g["rUpPu"]
    gensSheet.cell(row=gItr+2, column=rDnColNum).value = g["rDnPu"]

# write data to generators onbar sheet
onbarSheetName = "DCOnbar"
# check - if onbar data sheet of gams input excel file exists
if not onbarSheetName in gamsExcel.sheetnames:
    print('Onbar data sheet does not exist in gams input excel file')
    exit(0)
onbarSheet = gamsExcel[onbarSheetName]

# fetch onbar data
schRepo = SchedulesRepo(dbHost, dbName, dbUname, dbPass)

# populate data to onbar sheet
for gItr, g in enumerate(gens):
    onbarSheet.cell(row=gItr+2, column=1).value = g["name"]
    onbarSheet.cell(row=gItr+2, column=2).value = 1

    genOnbarRows = schRepo.getGenSchedules(
        "onbar", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check - check if we got 96 rows for loading onbar data for a generator for the desired date
    if not len(genOnbarRows) == 96:
        print("96 rows not present in onbar data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    for blkItr in range(len(genOnbarRows)):
        onbarSheet.cell(row=gItr+2, column=blkItr +
                        3).value = genOnbarRows[blkItr]["schVal"]

# write data to generators schedule sheet
schSheetName = "Schedule"
# check - if schedule data sheet of gams input excel file exists
if not schSheetName in gamsExcel.sheetnames:
    print('schedule data sheet does not exist in gams input excel file')
    exit(0)
schSheet = gamsExcel[schSheetName]

# populate data to schedule sheet
for gItr, g in enumerate(gens):
    schSheet.cell(row=gItr+2, column=1).value = g["name"]
    schSheet.cell(row=gItr+2, column=2).value = 1

    genSchRows = schRepo.getGenSchedules(
        "sch", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check - check if we got 96 rows for loading schedule data for a generator for the desired date
    if not len(genSchRows) == 96:
        print("96 rows not present in schedule data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    for blkItr in range(len(genSchRows)):
        schSheet.cell(row=gItr+2, column=blkItr +
                      3).value = genSchRows[blkItr]["schVal"]

# save and close the gams input excel file
gamsExcel.save(gamsExcelPath)
gamsExcel.close()

# run GAMS on input excel file
isGamsExecSuccess = runGams(gamsExePath, gamsCodePath, gamsLstPath)
# check - check if gams execution is successful
if not isGamsExecSuccess:
    print("GAMS execution was not successful...")
    exit(0)

# push optimal schedule output data from GAMS excel to db
resultsSheetName = "Result Report"
# check - if gams excel results sheet of gams input excel file exists
if not (resultsSheetName in gamsExcel.sheetnames):
    print('GAMS results data sheet does not exist in gams excel file...')
    exit(0)

# read dataframe from output sheet
gamsResDf = pd.read_excel(gamsExcelPath, sheet_name=resultsSheetName)
gamsResCols = gamsResDf.columns.tolist()
# check - check if gams excel result sheet was populated with at least 98 blocks
if len(gamsResCols) < 98:
    print("result sheet does not have at least 98 columns...")
    exit(0)

gamsResCols = gamsResCols[0:98]

# check - check if the gams excel result sheet was populated with all 96 blocks
isAllBlksInResPresent = all([(x in gamsResCols) for x in range(1, 97)])
if not isAllBlksInResPresent:
    print("result sheet does not have columns named 1 to 96")
    exit(0)

# check - check if all the database generators are present in the gams excel result sheet
resGens = gamsResDf.iloc[:, 0].tolist()
dbGenNames = [g["name"] for g in gens]
isAllDbGensInRes = all([(x in resGens) for x in dbGenNames])
if not isAllDbGensInRes:
    print("All DB generators are not in GAMS result sheet")
    exit(0)

# create generator names dictionary for ids lookup
genIdsDict = {}
for g in gens:
    genIdsDict[g["name"]] = g["id"]

# create optimal schedules rows
optSchRows = []
for genItr in range(gamsResDf.shape[0]):
    genName = gamsResDf.iloc[genItr, 0]
    if not genName in genIdsDict:
        continue
    genId = genIdsDict[genName]
    for blk in range(1, 97):
        schVal = gamsResDf[blk].iloc[genItr]
        schRow: ISchRow = {
            "schTime": targetDt+dt.timedelta(minutes=15*(blk-1)),
            "genId": genId,
            "schType": "opt",
            "schVal": schVal,
            "rev": 0
        }
        optSchRows.append(schRow)

# push optimal schedules data to db
isSchInsertSuccess = schRepo.insertSchedules(optSchRows)
print("Optimal Schedule Insertion status = {0}".format(isSchInsertSuccess))

# push SMP output data from GAMS excel to db
# read dataframe from output sheet
smpResultSheet = "Results2"
smpResDf = pd.read_excel(gamsExcelPath, sheet_name=smpResultSheet,
                         skiprows=5, nrows=1, usecols=range(27, 27+96))
smpResCols = smpResDf.columns.tolist()
try:
    smpResCols = [int(x) for x in smpResCols]
    smpResDf.columns = smpResCols
except:
    print("unable get column names as number in smp data parsing...")
    exit(0)

# check if all numbers from 1 to 96 are present in the smp dataframe
if not all([x in smpResCols for x in range(1, 97)]):
    print("96 blocks not present in smp data columns...")
    exit(0)

smpRows: List[ISmpRow] = []
for blkItr in range(1, 97):
    smpRow: ISmpRow = {
        "dataTime": targetDt + dt.timedelta(minutes=(blkItr-1)*15),
        "regTag": 'g',
        "smpVal": smpResDf[blkItr].iloc[0].item(),
        "rev": 0
    }
    smpRows.append(smpRow)
# push optimal schedules data to db
smpRepo = SmpRepo(dbHost, dbName, dbUname, dbPass)
isSmpInsertSuccess = smpRepo.insertSmp(smpRows)
print("SMP Insertion status = {0}".format(isSchInsertSuccess))

print("execution complete...")
