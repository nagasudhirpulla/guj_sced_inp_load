import argparse
import datetime as dt
import os
from typing import List

import pandas as pd

from src.config.appConfig import loadAppConfig
from src.repos.gensMasterDataRepo import GensMasterRepo
from src.repos.schDataRepo import SchedulesRepo
from src.typeDefs.schRow import ISchRow

import openpyxl

# read config file
appConf = loadAppConfig()
dbHost = appConf["dbHost"]
dbName = appConf["dbName"]
dbUname = appConf["dbUname"]
dbPass = appConf["dbPass"]

# make target date as tomorrow
targetDt = dt.datetime.now() + dt.timedelta(days=1)
targetDt = dt.datetime(targetDt.year, targetDt.month, targetDt.day)

# get target date from command line if present
parser = argparse.ArgumentParser()
parser.add_argument('--date', help='target Date')
args = parser.parse_args()
targetDtStr = args.date
if not targetDtStr == None:
    targetDt = dt.datetime.strptime(targetDtStr, "%Y-%m-%d")

# derive the filenames for the target date
gamsExcelPath = appConf["gamsExcelPath"]
# check if gams excel file exists
isGamsExcelPresent = os.path.isfile(gamsExcelPath)

if not isGamsExcelPresent:
    print("GAMS input excel not present...")
    exit(0)

# get the generators info from db
gensRepo = GensMasterRepo(
    appConf["dbHost"], appConf["dbName"], appConf["dbUname"], appConf["dbPass"])
gens = gensRepo.getGens()

gamsExcel = openpyxl.load_workbook(gamsExcelPath)

# write data to generators data sheet
# Data sheet column numbers
gensSheetName = "Data"
stationColNum = 3
vcColNum = 8
unitCapColNum = 11
tmColNum = 12
rUpColNum = 13
rDnColNum = 14
if gensSheetName in gamsExcel.sheetnames:
    print('Generators data sheet does not exist in gams input excel file')
gensSheet = gamsExcel[gensSheetName]
for gItr, g in enumerate(gens):
    gensSheet.cell(row=gItr+2, column=stationColNum).value = g["name"]
    gensSheet.cell(row=gItr+2, column=vcColNum).value = g["vcPu"]
    gensSheet.cell(row=gItr+2, column=unitCapColNum).value = g["capPu"]
    gensSheet.cell(row=gItr+2, column=tmColNum).value = g["tmPu"]
    gensSheet.cell(row=gItr+2, column=rUpColNum).value = g["rUpPu"]
    gensSheet.cell(row=gItr+2, column=rDnColNum).value = g["rDnPu"]

# write data to generators onbar sheet
onbarSheetName = "DCOnbar"
if not onbarSheetName in gamsExcel.sheetnames:
    print('Onbar data sheet does not exist in gams input excel file')
    exit(0)
onbarSheet = gamsExcel[onbarSheetName]

# fetch onbar data
schRepo = SchedulesRepo(dbHost, dbName, dbUname, dbPass)

# populate data to onbar sheet
for gItr, g in enumerate(gens):
    onbarSheet.cell(row=gItr+2, column=1).value = g["name"]
    onbarSheet.cell(row=gItr+2, column=2).value = 1

    genOnbarRows = schRepo.getGenSchedules(
        "onbar", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check if we got 96 rows
    if not len(genOnbarRows) == 96:
        print("96 rows not present in onabar data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    for blkItr in range(len(genOnbarRows)):
        onbarSheet.cell(row=gItr+2, column=blkItr +
                        3).value = genOnbarRows[blkItr]["schVal"]

# write data to generators onbar sheet
schSheetName = "Schedule"
if not schSheetName in gamsExcel.sheetnames:
    print('Onbar data sheet does not exist in gams input excel file')
    exit(0)
schSheet = gamsExcel[schSheetName]

# populate data to onbar sheet
for gItr, g in enumerate(gens):
    schSheet.cell(row=gItr+2, column=1).value = g["name"]
    schSheet.cell(row=gItr+2, column=2).value = 1

    genSchRows = schRepo.getGenSchedules(
        "sch", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check if we got 96 rows
    if not len(genSchRows) == 96:
        print("96 rows not present in schedule data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    for blkItr in range(len(genSchRows)):
        schSheet.cell(row=gItr+2, column=blkItr +
                      3).value = genSchRows[blkItr]["schVal"]

gamsExcel.save(gamsExcelPath)

print("execution complete...")
