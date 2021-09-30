import argparse
import datetime as dt
import math
import os

import pandas as pd
from openpyxl import Workbook

from src.config.appConfig import loadAppConfig
from src.repos.gensMasterDataRepo import GensMasterRepo
from src.repos.schDataRepo import SchedulesRepo
from src.services.ftpService import uploadFileToFtp

# read config file
appConf = loadAppConfig()
dbHost = appConf["dbHost"]
dbName = appConf["dbName"]
dbUname = appConf["dbUname"]
dbPass = appConf["dbPass"]
gamsExePath = appConf["gamsExePath"]
gamsCodePath = appConf["gamsCodePath"]
gamsLstPath = appConf["gamsLstPath"]
gamsExcelPath = appConf["gamsExcelPath"]
ftpHost = appConf["ftpHost"]
ftpUname = appConf["ftpUname"]
ftpPass = appConf["ftpPass"]
ftpResFolder = appConf["ftpResultsFolder"]

# make target date as tomorrow
targetDt = dt.datetime.now() + dt.timedelta(days=1)
targetDt = dt.datetime(targetDt.year, targetDt.month, targetDt.day)

# get target date from command line if present
parser = argparse.ArgumentParser()
parser.add_argument('--date', help='target Date')
parser.add_argument('--noftp', action="store_true")
args = parser.parse_args()
targetDtStr = args.date
noFtp = args.noftp
if not targetDtStr == None:
    targetDt = dt.datetime.strptime(targetDtStr, "%Y-%m-%d")

resDumpFolder = appConf["resultsDumpFolder"]
# check - check if results dumping folder exists
if not os.path.isdir(resDumpFolder):
    print("results dumping folder doesnot exist...")
    exit(0)

# create workbook object to dump excel data
wb = Workbook()

# write generators master data
gensSheet = wb.active
gensSheet.title = "Data"

# create onbar sheet
onbarSheet = wb.create_sheet("DCOnbar")
# populate header to onbar sheet
onbarSheet.cell(row=1, column=1).value = "Plant Name"
onbarSheet.cell(row=1, column=2).value = "Region"
for blk in range(1, 97):
    onbarSheet.cell(row=1, column=blk+2).value = blk

# create schedule sheet
schSheet = wb.create_sheet("Schedule")
# populate header to schedules sheet
schSheet.cell(row=1, column=1).value = "Plant Name"
schSheet.cell(row=1, column=2).value = "Region"
for blk in range(1, 97):
    schSheet.cell(row=1, column=blk+2).value = blk

# create Optimal Schedule sheet
optSheet = wb.create_sheet("Optimal Schedule")
# populate header to optimal schedules sheet
optSheet.cell(row=1, column=1).value = "Plant Name"
optSheet.cell(row=1, column=2).value = "Region"
for blk in range(1, 97):
    optSheet.cell(row=1, column=blk+2).value = blk

# create SCED sheet
scedSheet = wb.create_sheet("SCED")
# populate header to sced sheet
scedSheet.cell(row=1, column=1).value = "Plant Name"
scedSheet.cell(row=1, column=2).value = "Region"
for blk in range(1, 97):
    scedSheet.cell(row=1, column=blk+2).value = blk

# create Number of units sheet
numUnitsSheet = wb.create_sheet("NumUnits")
# populate header to number of units sheet
numUnitsSheet.cell(row=1, column=1).value = "Plant Name"
numUnitsSheet.cell(row=1, column=2).value = "Region"
for blk in range(1, 97):
    numUnitsSheet.cell(row=1, column=blk+2).value = blk

# create cost sheet
costSheet = wb.create_sheet("ScedCost")
# populate header to number of units sheet
costSheet.cell(row=1, column=1).value = "Plant Name"
costSheet.cell(row=1, column=2).value = "Region"
for blk in range(1, 97):
    costSheet.cell(row=1, column=blk+2).value = blk

# create summary sheet
summarySheet = wb.create_sheet("Summary")
# populate header to number of units sheet
summarySheet.cell(row=1, column=1).value = "Plant Name"
summarySheet.cell(row=1, column=2).value = "Region"
for hItr, headr in enumerate(["Plant Name", "VC Paise/MWH", "Max. On-bar (MW)", "On-bar Energy (MWH)", "Max. Inj. Schedule (MW)",
                              "Inj. Schedule Energy (MWH)", "Up Reserve (MWH)", "Down Reserve (MWH)", "Max. Optimal Schedule (MW)",
                              "Optimal Schedule Energy (MWH)", "Max SCED (MW)", "Min SCED (MW)", "SCED Energy (MWH)",
                              "Cost Incurred (Lakhs)", "Cost Savings (Lakhs)", "Net Savings (Lakhs)"]):
    summarySheet.cell(row=1, column=hItr+1).value = headr

# Initialize summary sheet all generators row values
dayOnbarMwh = 0
daySchMwh = 0
dayOptMwh = 0
dayTmMwh = 0
dayScedMwh = 0
dayScedCost = 0
dayScedSaving = 0

# get the generators info from db
gensRepo = GensMasterRepo(
    dbHost, dbName, dbUname, dbPass)
gens = gensRepo.getGens()

stationColNum = 3
vcColNum = 8
unitCapColNum = 11
tmColNum = 12
rUpColNum = 13
rDnColNum = 14
# populate header for 1st row
gensSheet.cell(row=1, column=stationColNum).value = "Plant Name"
gensSheet.cell(row=1, column=vcColNum).value = "Variable Cost per unit"
gensSheet.cell(
    row=1, column=unitCapColNum).value = "Avg. Unit capacity"
gensSheet.cell(row=1, column=tmColNum).value = "Tech Min Per Unit"
gensSheet.cell(row=1, column=rUpColNum).value = "Ramp Up per unit"
gensSheet.cell(row=1, column=rDnColNum).value = "Ramp Dn per unit"
for gItr, g in enumerate(gens):
    # populate generator data
    gensSheet.cell(row=gItr+2, column=stationColNum).value = g["name"]
    gensSheet.cell(row=gItr+2, column=vcColNum).value = g["vcPu"]
    gensSheet.cell(row=gItr+2, column=unitCapColNum).value = g["capPu"]
    gensSheet.cell(row=gItr+2, column=tmColNum).value = g["tmPu"]
    gensSheet.cell(row=gItr+2, column=rUpColNum).value = g["rUpPu"]
    gensSheet.cell(row=gItr+2, column=rDnColNum).value = g["rDnPu"]

# fetch onbar, sch, sced, optimal schedule data
schRepo = SchedulesRepo(dbHost, dbName, dbUname, dbPass)

# populate data to excel sheets
for gItr, g in enumerate(gens):
    onbarSheet.cell(row=gItr+2, column=1).value = g["name"]
    onbarSheet.cell(row=gItr+2, column=2).value = 1

    genOnbarRows = schRepo.getGenSchedules(
        "onbar", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check - check if we got 96 rows for loading onbar data for a generator for the desired date
    if not len(genOnbarRows) == 96:
        print("96 rows not present in onbar data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    genMaxOnbar = 0
    genOnbarMwh = 0
    genTmMwh = 0
    for blkItr in range(len(genOnbarRows)):
        onbarVal = genOnbarRows[blkItr]["schVal"]
        tmVal = g["tmPu"]*math.ceil(onbarVal*0.95/g["capPu"])
        if onbarVal > genMaxOnbar:
            genMaxOnbar = onbarVal
        genOnbarMwh += onbarVal
        genTmMwh += tmVal
        onbarSheet.cell(row=gItr+2, column=blkItr + 3).value = onbarVal
    genOnbarMwh /= 4
    genTmMwh /= 4

    # populate schedule data to Schedule sheet
    schSheet.cell(row=gItr+2, column=1).value = g["name"]
    schSheet.cell(row=gItr+2, column=2).value = 1

    genSchRows = schRepo.getGenSchedules(
        "sch", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check - check if we got 96 rows for loading sced data for a generator for the desired date
    if not len(genSchRows) == 96:
        print("96 rows not present in schedule data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    genMaxSch = 0
    genSchMwh = 0
    for blkItr in range(len(genSchRows)):
        schVal = genSchRows[blkItr]["schVal"]
        if schVal > genMaxSch:
            genMaxSch = schVal
        genSchMwh += schVal
        schSheet.cell(row=gItr+2, column=blkItr + 3).value = schVal
    genSchMwh /= 4

    # populate data to optimal schedule sheet
    optSheet.cell(row=gItr+2, column=1).value = g["name"]
    optSheet.cell(row=gItr+2, column=2).value = 1

    genOptRows = schRepo.getGenSchedules(
        "opt", g["id"], 0, targetDt, targetDt+dt.timedelta(hours=23, minutes=59))
    # check - check if we got 96 rows for loading sced data for a generator for the desired date
    if not len(genOptRows) == 96:
        print("96 rows not present in optimal schedule data of {0} for the date {1}".format(
            g["name"], targetDt))
        exit(0)

    genMaxOpt = 0
    genOptMwh = 0
    for blkItr in range(len(genOptRows)):
        optVal = genOptRows[blkItr]["schVal"]
        if optVal > genMaxOpt:
            genMaxOpt = optVal
        genOptMwh += optVal
        optSheet.cell(row=gItr+2, column=blkItr + 3).value = optVal
    genOptMwh /= 4

    # populate data to sced sheet, number of units sheet, cost sheet and summary
    scedSheet.cell(row=gItr+2, column=1).value = g["name"]
    scedSheet.cell(row=gItr+2, column=2).value = 1

    numUnitsSheet.cell(row=gItr+2, column=1).value = g["name"]
    numUnitsSheet.cell(row=gItr+2, column=2).value = 1

    costSheet.cell(row=gItr+2, column=1).value = g["name"]
    costSheet.cell(row=gItr+2, column=2).value = 1

    summarySheet.cell(row=gItr+2, column=1).value = g["name"]
    summarySheet.cell(row=gItr+2, column=2).value = g["vcPu"]

    # check - check if schedule, Onbar and optimal schedule have same number of rows
    if not (len(genOptRows) == len(genSchRows) == len(genOnbarRows)):
        print("Schedule, Onbar and optimal schedule rows are not of same size for {0}".format(
            targetDt))
        exit(0)
    genMaxSced = None
    genMinSced = None
    genScedMwh = 0
    genScedCost = 0
    genScedSaving = 0
    for blkItr in range(len(genOptRows)):
        scedVal = genOptRows[blkItr]["schVal"] - genSchRows[blkItr]["schVal"]
        if blkItr == 0:
            genMaxSced = scedVal
            genMinSced = scedVal
        else:
            if scedVal > genMaxSced:
                genMaxSced = scedVal
            if scedVal < genMinSced:
                genMinSced = scedVal
        genScedMwh += scedVal
        scedSheet.cell(row=gItr+2, column=blkItr +
                       3).value = scedVal

        numUnitsVal = math.ceil(
            0.95*genOnbarRows[blkItr]["schVal"]/g["capPu"])
        numUnitsSheet.cell(row=gItr+2, column=blkItr +
                           3).value = numUnitsVal
        genVcPu = g["vcPu"]
        scedBlkCost = scedVal*genVcPu*-2.5
        if scedBlkCost < 0:
            genScedCost += scedBlkCost
        else:
            genScedSaving += scedBlkCost
        costSheet.cell(row=gItr+2, column=blkItr + 3).value = scedBlkCost
    genScedMwh /= 4
    genScedCost /= 100000
    genScedSaving /= 100000

    dayOnbarMwh += genOnbarMwh
    daySchMwh += genSchMwh
    dayOptMwh += genOptMwh
    dayScedMwh += genScedMwh
    dayScedCost += genScedCost
    dayScedSaving += genScedSaving
    dayTmMwh += genTmMwh

    summarySheet.cell(row=gItr+2, column=3).value = genMaxOnbar
    summarySheet.cell(row=gItr+2, column=4).value = genOnbarMwh
    summarySheet.cell(row=gItr+2, column=5).value = genMaxSch
    summarySheet.cell(row=gItr+2, column=6).value = genSchMwh
    summarySheet.cell(row=gItr+2, column=7).value = genOnbarMwh-genSchMwh
    summarySheet.cell(row=gItr+2, column=8).value = genSchMwh-genTmMwh
    summarySheet.cell(row=gItr+2, column=9).value = genMaxOpt
    summarySheet.cell(row=gItr+2, column=10).value = genOptMwh
    summarySheet.cell(row=gItr+2, column=11).value = genMaxSced
    summarySheet.cell(row=gItr+2, column=12).value = genMinSced
    summarySheet.cell(row=gItr+2, column=13).value = genScedMwh
    summarySheet.cell(row=gItr+2, column=14).value = genScedCost
    summarySheet.cell(row=gItr+2, column=15).value = genScedSaving
    summarySheet.cell(
        row=gItr+2, column=16).value = genScedSaving+genScedCost

# populate total generators row values after all the generators in summary sheet
summarySheet.cell(row=len(gens)+2, column=1).value = "TOTAL"
summarySheet.cell(row=len(gens)+2, column=4).value = dayOnbarMwh
summarySheet.cell(row=len(gens)+2, column=6).value = daySchMwh
summarySheet.cell(row=len(gens)+2, column=7).value = dayOnbarMwh-daySchMwh
summarySheet.cell(row=len(gens)+2, column=8).value = daySchMwh-dayTmMwh
summarySheet.cell(row=len(gens)+2, column=10).value = dayOptMwh
summarySheet.cell(row=len(gens)+2, column=13).value = dayScedMwh
summarySheet.cell(row=len(gens)+2, column=14).value = dayScedCost
summarySheet.cell(row=len(gens)+2, column=15).value = dayScedSaving
summarySheet.cell(
    row=len(gens)+2, column=16).value = dayScedSaving+dayScedCost


# derive excel filename and file path
resultsFilename = "sced_results_{0}.xlsx".format(
    dt.datetime.strftime(targetDt, "%Y_%m_%d"))
resultsFilePath = os.path.join(resDumpFolder, resultsFilename)

# save workbook in dumps folder
wb.save(resultsFilePath)
wb.close()

# copy results file to ftp location
if not noFtp:
    isResFtpUploadSuccess = uploadFileToFtp(
        resultsFilePath, ftpHost, ftpUname, ftpPass, ftpResFolder)
    print("Results excel FTP upload status = {0}".format(
        isResFtpUploadSuccess))

print("SCED output data excel publish program complete...")
